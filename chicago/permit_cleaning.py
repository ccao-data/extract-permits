"""
Chicago Permit Ingest Process - Automation

This script automates the current process for cleaning permit data from the Chicago Data Portal's Building Permits table
and preparing it for upload to iasWorld via SmartFile. This involves fetching the data, cleaning up certain fields,
organizing columns to match the SmartFile template, and batching the data into Excel workbooks of 200 rows each. This process
also splits off data that is ready for upload from data that still needs some manual review before upload, saving each
in separate Excel workbooks in separate folders. Data that need review are split into two categories and corresponding folders/files:
those with quick fixes for fields over character or amount limits, and those with more complicated fixes for missing and/or invalid fields.

The following optional environment variables can be set:
    AWS_ATHENA_S3_STAGING_DIR: S3 path where Athena query results are stored
    AWS_REGION: Region that AWS operations should be run in

The script also expects three positional arguments:
    * start_date (str, YYYY-MM-DD): The lower bound date to use for filtering permits
    * end_date (str, YYYY-MM-DD): The upper bound date to use for filtering
    * deduplicate (bool): Whether to filter out permits that already exist in iasworld
"""

import decimal
import math
import os
import re
import sys
from datetime import datetime

import numpy as np
import openpyxl
import openpyxl.styles
import pandas as pd
import requests
from openpyxl.utils import get_column_letter
from pyathena import connect
from pyathena.cursor import Cursor
from pyathena.pandas.util import as_pandas


def parse_args() -> tuple[str, str, bool]:
    """Helper function to parse and validate command line args to this
    script"""
    if len(sys.argv) < 4:
        print(
            "Usage: permit_cleaning.py <start_date> <end_date> <deduplicate>"
        )
        sys.exit(1)

    start_date_str, end_date_str, deduplicate = (
        sys.argv[1],
        sys.argv[2],
        sys.argv[3],
    )

    try:
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
    except ValueError:
        print(
            f"Invalid start_date format: '{start_date_str}'. Expected YYYY-MM-DD."
        )
        sys.exit(1)

    try:
        end_date = datetime.strptime(end_date_str, "%Y-%m-%d")
    except ValueError:
        print(
            f"Invalid end_date format: '{end_date_str}'. Expected YYYY-MM-DD."
        )
        sys.exit(1)

    if end_date < start_date:
        print("Error: end_date must be later than or equal to start_date.")
        sys.exit(1)

    deduplicate = deduplicate.lower() == "true"

    return start_date_str, end_date_str, deduplicate


def get_current_assessment_year(cursor: Cursor) -> str:
    """Query Athena for the current assessment year"""
    # Use PARDAT as the source of truth for the current assessment year.
    # This choice is somewhat arbitrary, since we could theoretically use
    # any table as the source of truth for this info, but PARDAT feels
    # relevant given that it is the base for our PIN universe and the
    # main purpose of this year is to help us construct that universe
    cursor.execute("""
        SELECT MAX(taxyr)
        FROM iasworld.pardat
        WHERE cur = 'Y'
            AND deactivat IS NULL
    """)
    return cursor.fetchall()[0][0]


def get_pin_cache_filename(year: str) -> str:
    """Given a year, return the name of a file that we can use to cache
    distinct PINs in that year"""
    return f"chicago_pin_universe_{year}.csv"


def pull_existing_pins_from_athena(cursor: Cursor, year: str) -> pd.DataFrame:
    """Connect to Athena and download all PINs in Chicago for a given year,
    saving the resulting data to a cache file according to the year."""
    SQL_QUERY = """
    SELECT
        CAST(u.pin AS varchar) AS pin,
        CAST(u.pin10 AS varchar) AS pin10,
        a.prop_address_full
    FROM default.vw_pin_universe u
    LEFT JOIN default.vw_pin_address a
        ON u.pin = a.pin
        AND u.year = a.year
    WHERE u.triad_name = 'City'
    AND u.year = %(year)s
    """
    cursor.execute(SQL_QUERY, {"year": year})
    chicago_pin_universe = as_pandas(cursor)
    pin_cache_filename = get_pin_cache_filename(year)
    chicago_pin_universe.to_csv(pin_cache_filename, index=False)

    return chicago_pin_universe


def download_permits(start_date: str, end_date: str) -> pd.DataFrame:
    """Download permits from the Chicago open data portal in the dataframe
    with issue dates between `start_date` and `end_date`"""
    params = {
        # Assume we've already validated the start and end date strings for
        # YYYY-MM-DD format
        "$where": f"issue_date between '{start_date}' and '{end_date}'",
        "$order": "issue_date DESC",
        "$limit": 10000000,  # Artificial limit to override the default
    }
    url = "https://data.cityofchicago.org/resource/ydr8-5enu.json"
    permits_response = requests.get(url, params=params)
    permits_response.raise_for_status()
    permits = permits_response.json()
    permits_df = pd.DataFrame(permits)
    return permits_df


def expand_multi_pin_permits(df):
    """
    Data from the Chicago open data permits table (data this script works with) has rows uniquely identified by permit number.
    Permits can apply to multiple PINs, in which case the PIN_LIST column will
    be a pipe-separated value representing multiple PINs.
    We want rows that are uniquely identified by PIN and permit number.
    This function creates new rows for each PIN in multi-PIN permits and saves the relevant PIN in pin_solo.
    """

    def parse_pin_list(pin_list_str):
        """Parse the PIN_LIST column. The column is formatted like so:

        * If no PINs are in the list, the value is NA, and we want to keep it
        * Otherwise, PINs are stored as pipe-separated values, and we want to
          parse them as lists so that we can pivot them out to one
          permit per PIN
        """
        if pd.isna(pin_list_str):
            return np.nan
        elif " | " in pin_list_str:
            pin_list = pin_list_str.split(" | ")
            # Remove duplicates while maintaining list order
            return list(dict.fromkeys(pin_list))
        else:
            return [pin_list_str]

    df["pin_list"] = df["pin_list"].apply(parse_pin_list)

    # Retain rows where pin_list is NA, since the pivot operation will
    # remove them but we want to keep them
    na_rows = df[df["pin_list"].isna()]
    df = df.dropna(subset=["pin_list"])

    # Pivot the dataframe longer by "pin_list" so that we have one row
    # per PIN with "solo_pin" being the PIN column
    df = (
        df.explode("pin_list")
        .reset_index(drop=True)
        .rename(columns={"pin_list": "solo_pin"})
    )

    # Add a column to track the position of the PIN in the PIN list for
    # ordering, starting at 1
    df["pin_type"] = df.groupby("permit_").cumcount() + 1
    df["pin_type"] = "pin" + df["pin_type"].astype(str)

    # Add back the NA rows
    df = pd.concat([df, na_rows], ignore_index=True)

    # Sort by the permit number, then the position of the PIN in the list,
    # so that PINs will be in order of their permit number and list position
    # in the output table
    df = df.sort_values(by=["permit_", "pin_type"]).reset_index(drop=True)

    return df


# update pin to match formatting of iasWorld
def format_pin(df):
    # iasWorld format doesn't include dashes
    df["pin_final"] = df["solo_pin"].astype("string").str.replace("-", "")

    # add zeros to 10-digit PINs to transform into 14-digits PINs
    def pad_pin(pin):
        if not pd.isna(pin):
            if len(pin) == 10:
                return pin + "0000"
            else:
                return pin
        else:
            return ""

    df["pin_final"] = df["pin_final"].apply(pad_pin)
    df["NOTE: 0000 added to PIN?"] = df["pin_final"].apply(
        lambda x: "Yes" if len(x) == 10 else "No"
    )
    return df


# Eliminate columns not included in permit upload and rename and order to match Smartfile excel format
def organize_columns(df):
    address_columns = ["street_number", "street_direction", "street_name"]
    df[address_columns] = df[address_columns].fillna("")
    df["Address"] = df[address_columns].astype("string").agg(" ".join, axis=1)

    df["issue_date"] = pd.to_datetime(
        df["issue_date"], format="%Y-%m-%dT%H:%M:%S.%f", errors="coerce"
    ).dt.strftime("%-m/%-d/%Y")

    column_renaming_dict = {
        "solo_pin": "Original PIN",
        "pin_final": "PIN* [PARID]",
        "permit_": "Local Permit No.* [USER28]",
        "issue_date": "Issue Date* [PERMDT]",
        "reported_cost": "Amount* [AMOUNT]",
        "Address": "Applicant Street Address* [ADDR1]",
        "city_state": "Applicant City, State, Zip* [ADDR3]",
        "contact_1_name": "Applicant* [USER21]",
        "work_description": "Notes [NOTE1]",
    }

    data_relevant = df[
        [col for col in df.columns if col in column_renaming_dict]
    ]
    data_renamed = data_relevant.rename(columns=column_renaming_dict)

    column_order = [
        "Original PIN",  # will keep original PIN column for rows flagged for invalid PINs
        "PIN* [PARID]",
        "Local Permit No.* [USER28]",
        "Issue Date* [PERMDT]",
        "Desc 1* [DESC1]",
        "Desc 2 Code 1 [USER6]",
        "Desc 2 Code 2 [USER7]",
        "Desc 2 Code 3 [USER8]",
        "Amount* [AMOUNT]",
        "Assessable [IS_ASSESS]",
        "Applicant Street Address* [ADDR1]",
        "Applicant Address 2 [ADDR2]",
        "Applicant City, State, Zip* [ADDR3]",
        "Contact Phone* [PHONE]",
        "Applicant* [USER21]",
        "Notes [NOTE1]",
        "Occupy Dt [UDATE1]",
        "Submit Dt* [CERTDATE]",
        "Est Comp Dt [UDATE2]",
    ]

    data_all_cols = data_renamed.assign(
        **{col: None for col in column_order if col not in data_renamed}
    )
    data_ordered = data_all_cols[column_order]

    return data_ordered


# flag invalid PINs for review by analysts
def flag_invalid_pins(df, chicago_pin_universe):
    df["FLAG COMMENTS"] = ""

    # invalid 14-digit PIN flag
    df["FLAG, INVALID: PIN* [PARID]"] = np.where(
        df["PIN* [PARID]"] == "",
        0,
        ~df["PIN* [PARID]"].isin(chicago_pin_universe["pin"]),
    )

    # also check if 10-digit PINs are valid to narrow down on problematic portion of invalid PINs
    df["pin_10digit"] = df["PIN* [PARID]"].astype("string").str[:10]
    df["FLAG, INVALID: pin_10digit"] = np.where(
        df["pin_10digit"] == "",
        0,
        ~df["pin_10digit"].isin(chicago_pin_universe["pin10"]),
    )

    # create variable that is the numbers following the 10-digit PIN
    # (not pulling last 4 digits from the end in case there are PINs that are not 14-digits in Chicago permit data)
    df["pin_suffix"] = df["PIN* [PARID]"].astype("string").str[10:]

    # comment for rows with invalid PINs
    df["FLAG COMMENTS"] += df.apply(
        lambda row: "First 10 digits of PIN* [PARID] do not match a valid PIN10; "
        if row["FLAG, INVALID: pin_10digit"] == 1
        else (
            "First 10 digits of PIN* [PARID] match a valid PIN10, but last 4 digits do not match; "
            if row["FLAG, INVALID: PIN* [PARID]"] == 1
            else ""
        ),
        axis=1,
    )

    return df


def flag_fix_long_fields(df):
    # will use these abbreviations to shorten applicant name field (Applicant* [USER21]) within 50 character field limit
    name_shortening_dict = {
        "ASSOCIATION": "ASSN",
        "COMPANY": "CO",
        "BUILDING": "BLDG",
        "FOUNDATION": "FNDN",
        "ILLINOIS": "IL",
        "STREET": "ST",
        "BOULEVARD": "BLVD",
        "AVENUE": "AVE",
        "APARTMENT": "APT",
        "APARTMENTS": "APTS",
        "MANAGEMENT": "MGMT",
        "CORPORATION": "CORP",
        "INCORPORATED": "INC",
        "LIMITED": "LTD",
        "PLAZA": "PLZ",
    }

    df["Applicant* [USER21]"] = df["Applicant* [USER21]"].replace(
        name_shortening_dict, regex=True
    )

    # these fields have the following character limits in Smartfile / iasWorld, flag if over limit
    long_fields_to_flag = [
        (
            "FLAG, LENGTH: Applicant Name",
            "Applicant* [USER21]",
            50,
            "Applicant* [USER21] over 50 char limit by ",
        ),
        (
            "FLAG, LENGTH: Permit Number",
            "Local Permit No.* [USER28]",
            18,
            "Local Permit No.* [USER28] over 18 char limit by ",
        ),
        (
            "FLAG, LENGTH: Applicant Street Address",
            "Applicant Street Address* [ADDR1]",
            40,
            "Applicant Street Address* [ADDR1] over 40 char limit by ",
        ),
        (
            "FLAG, LENGTH: Note1",
            "Notes [NOTE1]",
            2000,
            "Notes [NOTE1] over 2000 char limit by ",
        ),
    ]

    for flag_name, column, limit, comment in long_fields_to_flag:
        df[flag_name] = df[column].apply(
            lambda val: 0
            if pd.isna(val)
            else (1 if len(str(val)) > limit else 0)
        )
        df["FLAG COMMENTS"] += df[column].apply(
            lambda val: ""
            if pd.isna(val)
            else (
                ""
                if len(str(val)) <= limit
                else comment + str(len(str(val)) - limit) + "; "
            )
        )

    # round Amount to closest dollar because smart file doesn't accept decimal amounts, then flag values above upper limit
    df["Amount* [AMOUNT]"] = (
        pd.to_numeric(df["Amount* [AMOUNT]"], errors="coerce")
        .round()
        .astype("Int64")
    )
    df["FLAG, VALUE: Amount"] = df["Amount* [AMOUNT]"].apply(
        lambda value: 0 if pd.isna(value) or value <= 2147483647 else 1
    )
    df["FLAG COMMENTS"] += df["Amount* [AMOUNT]"].apply(
        lambda value: ""
        if pd.isna(value) or value <= 2147483647
        else "Amount* [AMOUNT] over value limit of 2147483647; "
    )

    # also flag rows where fields are blank for manual review (for fields we're populating in smartfile template)
    empty_fields_to_flag = [
        ("FLAG, EMPTY: PIN", "PIN* [PARID]"),
        ("FLAG, EMPTY: Issue Date", "Issue Date* [PERMDT]"),
        ("FLAG, EMPTY: Amount", "Amount* [AMOUNT]"),
        ("FLAG, EMPTY: Applicant", "Applicant* [USER21]"),
        (
            "FLAG, EMPTY: Applicant Street Address",
            "Applicant Street Address* [ADDR1]",
        ),
        ("FLAG, EMPTY: Permit Number", "Local Permit No.* [USER28]"),
        ("FLAG, EMPTY: Note1", "Notes [NOTE1]"),
    ]

    for flag_name, column in empty_fields_to_flag:
        comment = column + " is missing; "
        df[flag_name] = df[column].apply(
            lambda val: 1 if pd.isna(val) or str(val).strip() == "" else 0
        )
        df["FLAG COMMENTS"] += df[flag_name].apply(
            lambda val: "" if val == 0 else comment
        )

    # Create flags based on the pin errors and then those based on other errors.
    # These are sorted into separate pages
    flag_cols = df.filter(regex="^FLAG,").columns.tolist()
    pin_flag_cols = [c for c in flag_cols if "pin" in c.lower()]
    other_flag_cols = [c for c in flag_cols if "pin" not in c.lower()]

    # Sum pin values to create the pin flags vs other flags
    if pin_flag_cols:
        df["FLAGS, TOTAL - PIN"] = df[pin_flag_cols].sum(axis=1)
    else:
        df["FLAGS, TOTAL - PIN"] = 0

    if other_flag_cols:
        df["FLAGS, TOTAL - OTHER"] = df[other_flag_cols].sum(axis=1)
    else:
        df["FLAGS, TOTAL - OTHER"] = 0

    # for ease of analysts viewing, edits flag columns to read "Yes" when row is flagged and blank otherwise (easier than columns of 0s and 1s)
    df[pin_flag_cols] = df[pin_flag_cols].replace({0: "", 1: "Yes"})
    df[other_flag_cols] = df[other_flag_cols].replace({0: "", 1: "Yes"})

    return df


# List of keywords to identify likely assessable permits.
# This is produced via a document provided by Valuations
# and Data Integrity.
# Build was in the provided document
# but is a component of too many words (building)

keywords = [
    "Addition",
    "Elevator",
    "Window",
    "Construction",
    "Garage",
    "Roof",
    "Demolition",
    "HVAC",
    "Flatwork",
    "Expand",
    "Basement",
    "Alarm",
    "Fire",
    "Bathroom",
    "Solar",
    "New",
    "Attic",
    "Vacant",
    "Conversion",
    "Rehab",
    "Enclosed porch",
    "Alteration",
    "EFP",
    "ADU",
    "A.D.U.",
    "Coach",
    "Accessory",
    "Extension",
    "Dormer",
    "Erect",
    "Proposed",
    "Wreck",
    "Finish",
    "Rec Room",
    "Convert",
    "Recreation room",
    "Sun Room",
    "Season",
]


# join addresses and format columns
def add_address_link_and_suggested_pins(df, chicago_pin_universe):
    # Collapse multiple pins per address into a single comma-separated string
    pin_map = (
        chicago_pin_universe.groupby(["prop_address_full"])["pin"]
        .apply(lambda pins: ", ".join(pins.astype(str).unique()))
        .reset_index()
    )

    # Merge using the collapsed mapping
    df = df.merge(
        pin_map,
        left_on=["Applicant Street Address* [ADDR1]"],
        right_on=["prop_address_full"],
        how="left",
    )

    # Insert Property Address column right after the Applicant Street Address column
    df.insert(
        df.columns.get_loc("Applicant Street Address* [ADDR1]") + 1,
        "Property Address",
        df["Applicant Street Address* [ADDR1]"],
    )

    # Suggested PINs (replace NA with empty string)
    df = df.rename(columns={"pin": "Suggested PINs"})
    df["Suggested PINs"] = df["Suggested PINs"].fillna("")

    # Drop the prop_address_full column (no longer needed)
    df = df.drop(columns=["prop_address_full"])

    # Add hyperlink for the Property Address
    df["Property Address"] = df["Property Address"].apply(
        lambda addr: (
            f'=HYPERLINK("https://maps.cookcountyil.gov/cookviewer/?search={addr}", "{addr}")'
            if pd.notna(addr)
            else ""
        )
    )

    # This uses three techniques to add a suggested PIN. If there is no PIN, it will say "NO PIN FOUND".
    # If there is a single 14-digit PIN, it will be a hyperlink.
    # If there are more than one PINs, it will be a comma-separated list of PINs. This is both the
    # result of joining based on pin10 and the fact that multiple pins may have the same address.
    def make_pin_hyperlink(pin_str):
        if pd.isna(pin_str):
            return "NO PIN FOUND"

        digits = re.sub(r"\D", "", pin_str)
        if len(digits) == 14:
            return f'=HYPERLINK("https://www.cookcountyassessoril.gov/pin/{digits}", "{pin_str}")'

        # This will be a list of comma separated pins
        return pin_str

    # Apply
    df["Suggested PINs"] = df["Suggested PINs"].apply(make_pin_hyperlink)

    # Create a comma separated list of matched keywords. This is derived from
    # the list called keywords.
    df = df.assign(
        **{
            "Matched Keywords": df["Notes [NOTE1]"].apply(
                lambda note: ", ".join(
                    [kw for kw in keywords if kw.lower() in str(note).lower()]
                )
            )
        }
    )
    return df


def deduplicate_permits(cursor, df, start_date, end_date):
    cursor.execute(
        """
            SELECT
                parid,
                permdt,
                amount,
                note2,
                user21,
                user28,
                user43
            FROM iasworld.permit
            WHERE permdt BETWEEN %(start_date)s AND %(end_date)s
        """,
        {"start_date": start_date, "end_date": end_date},
    )
    existing_permits = as_pandas(cursor)
    workbook_to_iasworld_col_map = {
        "PIN* [PARID]": "parid",
        "Issue Date* [PERMDT]": "permdt",
        "Amount* [AMOUNT]": "amount",
        "Applicant Street Address* [ADDR1]": "note2",
        "Local Permit No.* [USER28]": "user28",
        "Notes [NOTE1]": "user43",
    }
    new_permits = df.copy()
    for workbook_key, iasworld_key in workbook_to_iasworld_col_map.items():
        new_permits[iasworld_key] = new_permits[workbook_key]

    # Transform new columns to ensure they match the iasworld formatting
    new_permits["amount"] = new_permits["amount"].apply(
        lambda x: decimal.Decimal("{:.2f}".format(x))
        if not pd.isnull(x)
        else x
    )

    new_permits["permdt"] = new_permits["permdt"].apply(
        lambda x: datetime.strptime(x, "%m/%d/%Y").strftime(
            "%Y-%m-%d %H:%M:%S.%f"
        )[:-3]
    )
    new_permits["note2"] = new_permits["note2"] + ",,CHICAGO, IL"
    new_permits["user43"] = (
        new_permits["user43"]
        # Replace special characters that Smartfile removes
        .str.replace(r"""[():;+#*&'"@½]""", "", regex=True)
        # Truncate description to match Smartfile length limit
        .str.slice(0, 259)
    )

    # Antijoin new_permits to existing_permits to find permits that do
    # not exist in iasworld
    merged_permits = pd.merge(
        new_permits,
        existing_permits,
        how="left",
        on=list(workbook_to_iasworld_col_map.values()),
        indicator=True,
    )
    true_new_permits = merged_permits[merged_permits["_merge"] == "left_only"]
    true_new_permits = true_new_permits.drop("_merge", axis=1)
    for iasworld_key in workbook_to_iasworld_col_map.values():
        true_new_permits = true_new_permits.drop(iasworld_key, axis=1)

    return true_new_permits


def gen_file_base_name(start_date, end_date):
    return f"{start_date}_to_{end_date}_permits_"


def _build_textjoin_errors_formula(row: int) -> str:
    """Return the TEXTJOIN formula for the Errors column at a given row"""
    r = row
    return (
        f'=_xlfn.TEXTJOIN(", ", TRUE,\n'
        f'  IF(LEN(TRIM(D{r}))=0, "Missing PIN14", ""),\n'
        f'  IF(COUNTIF(\'Universe of Valid PINs\'!A:A, D{r}) > 0, "", "Provide Valid Pin"),\n'
        f'  IF(LEN(TRIM(D{r}))<>14, "PIN is not 14 digits", ""),\n'
        f'  IF(LEN(R{r})>50, "Applicant Name > 50 characters", ""),\n'
        f'  IF(LEN(F{r})>40, "Address > 40 characters", ""),\n'
        f'  IF(LEN(S{r})>2000, "Work Description > 2000 characters", ""),\n'
        f'  IF(AND(ISNUMBER(M{r}), M{r}>2147483647), "Amount exceeds limit", ""),\n'
        f'  IF(OR(H{r}="", NOT(ISNUMBER(DATEVALUE(H{r})))), "Missing or Invalid Issue Date", ""),\n'
        f'  IF(OR(M{r}="", NOT(ISNUMBER(M{r}))), "Missing Amount", ""),\n'
        f'  IF(LEN(TRIM(R{r}))=0, "Missing Applicant", ""),\n'
        f'  IF(LEN(TRIM(F{r}))=0, "Missing Applicant Street Address", ""),\n'
        f'  IF(LEN(TRIM(G{r}))=0, "Missing Permit Number", ""),\n'
        f'  IF(LEN(TRIM(S{r}))=0, "Missing Work Description", "")\n'
        f")"
    )


# Column layout for the "Needs Review" sheet
REVIEW_HEADERS = [
    "Row Number",  # A - original row number (# [LLINE])
    "Errors",  # B - TEXTJOIN formula
    "Suggested PINs",  # C
    "PIN",  # D - PIN* [PARID]
    "Suggested Property Address",  # E - hyperlink to Cook County viewer
    "Applicant Street Address",  # F - Applicant Street Address* [ADDR1]
    "Local Permit No.",  # G - Local Permit No.* [USER28]
    "Issue Date",  # H - Issue Date* [PERMDT]
    "Desc 1* [DESC1]",  # I
    "Desc 2 Code 1 [USER6]",  # J
    "Desc 2 Code 2 [USER7]",  # K
    "Desc 2 Code 3 [USER8]",  # L
    "Amount",  # M - Amount* [AMOUNT]
    "Assessable [IS_ASSESS]",  # N
    "Applicant Address 2 [ADDR2]",  # O
    "Applicant City, State, Zip* [ADDR3]",  # P
    "Contact Phone* [PHONE]",  # Q
    "Applicant",  # R - Applicant* [USER21]
    "Notes",  # S - Notes [NOTE1]
    "Occupy Dt [UDATE1]",  # T
    "Submit Dt* [CERTDATE]",  # U
    "Est Comp Dt [UDATE2]",  # V
    "Matched Keywords",  # W
    "Errors are Resolved",  # X
]

REVIEW_HIDDEN_COLS = {9, 10, 11, 12, 14, 15, 17, 20, 21, 22}


def save_xlsx_files(df, max_rows, file_base_name):
    # Separate rows ready for upload from those needing review
    df_ready = df[
        (df["FLAGS, TOTAL - PIN"] == 0) & (df["FLAGS, TOTAL - OTHER"] == 0)
    ].reset_index()
    df_ready = df_ready.drop(
        columns=df_ready.filter(like="FLAG").columns
    ).drop(
        columns=[
            "index",
            "Original PIN",
            "pin_10digit",
            "pin_suffix",
            "Property Address",
            "Suggested PINs",
            "Matched Keywords",
        ]
    )

    # All rows with any flag go to a single "Needs Review" sheet
    df_needs_review = df[
        (df["FLAGS, TOTAL - PIN"] > 0) | (df["FLAGS, TOTAL - OTHER"] > 0)
    ].reset_index(drop=True)

    print("# rows ready for upload: ", len(df_ready))
    print("# rows needing review: ", len(df_needs_review))

    folder_for_files_ready = (
        datetime.today().date().strftime("files_for_smartfile_%Y_%m_%d")
    )
    os.makedirs(folder_for_files_ready, exist_ok=True)
    folder_for_files_review = (
        datetime.today().date().strftime("files_for_review_%Y_%m_%d")
    )
    os.makedirs(folder_for_files_review, exist_ok=True)

    # Save ready permits batched into max_rows per file
    num_files_ready = math.ceil(len(df_ready) / max_rows)
    print(f"Creating {num_files_ready} xlsx files ready for SmartFile upload")
    for i in range(num_files_ready):
        file_dataframe = df_ready.iloc[
            i * max_rows : (i + 1) * max_rows
        ].copy()
        file_dataframe.reset_index(drop=True, inplace=True)
        file_dataframe.index = file_dataframe.index + 1
        file_dataframe.index.name = "# [LLINE]"
        file_dataframe = file_dataframe.reset_index()
        file_name = os.path.join(
            folder_for_files_ready,
            file_base_name + f"ready_for_upload_{i + 1}.xlsx",
        )
        file_dataframe.to_excel(file_name, index=False, engine="xlsxwriter")

    # Build the review workbook with the new single-sheet structure
    file_name_review = os.path.join(
        folder_for_files_review, file_base_name + "needing_review.xlsm"
    )

    wb = openpyxl.Workbook()

    # --- "Needs Review" sheet ---
    ws_review = wb.active
    ws_review.title = "Needs Review"

    bold_font = openpyxl.styles.Font(bold=True, name="Arial")
    wrap_top = openpyxl.styles.Alignment(wrap_text=True, vertical="top")
    hyperlink_font = openpyxl.styles.Font(
        color="0000FF", underline="single", name="Arial"
    )
    normal_font = openpyxl.styles.Font(name="Arial")

    # Write header row
    for col_idx, header in enumerate(REVIEW_HEADERS, start=1):
        cell = ws_review.cell(row=1, column=col_idx, value=header)
        cell.font = bold_font

    # Map from our internal column names to the REVIEW_HEADERS column positions
    # REVIEW_HEADERS col positions (1-based):
    # A=1 Row Number, B=2 Errors(formula), C=3 Suggested PINs, D=4 PIN,
    # E=5 Suggested Property Address, F=6 Applicant Street Address,
    # G=7 Local Permit No., H=8 Issue Date, I=9 Desc1, J=10 D2C1, K=11 D2C2,
    # L=12 D2C3, M=13 Amount, N=14 Assessable, O=15 Addr2, P=16 City/State/Zip,
    # Q=17 Phone, R=18 Applicant, S=19 Notes, T=20 OccupyDt, U=21 SubmitDt,
    # V=22 EstCompDt, W=23 MatchedKeywords, X=24 ErrorsResolved

    col_map = {
        "PIN* [PARID]": 4,
        "Suggested PINs": 3,
        "Property Address": 5,  # hyperlink to Cook County viewer
        "Applicant Street Address* [ADDR1]": 6,
        "Local Permit No.* [USER28]": 7,
        "Issue Date* [PERMDT]": 8,
        "Desc 1* [DESC1]": 9,
        "Desc 2 Code 1 [USER6]": 10,
        "Desc 2 Code 2 [USER7]": 11,
        "Desc 2 Code 3 [USER8]": 12,
        "Amount* [AMOUNT]": 13,
        "Assessable [IS_ASSESS]": 14,
        "Applicant Address 2 [ADDR2]": 15,
        "Applicant City, State, Zip* [ADDR3]": 16,
        "Contact Phone* [PHONE]": 17,
        "Applicant* [USER21]": 18,
        "Notes [NOTE1]": 19,
        "Occupy Dt [UDATE1]": 20,
        "Submit Dt* [CERTDATE]": 21,
        "Est Comp Dt [UDATE2]": 22,
        "Matched Keywords": 23,
    }

    for data_row_idx, (_, row_data) in enumerate(
        df_needs_review.iterrows(), start=2
    ):
        # Col A: original row number (1-based index in the flagged set)
        ws_review.cell(
            row=data_row_idx, column=1, value=data_row_idx - 1
        ).font = normal_font

        # Col B: TEXTJOIN errors formula
        ws_review.cell(
            row=data_row_idx,
            column=2,
            value=_build_textjoin_errors_formula(data_row_idx),
        ).font = normal_font

        # Remaining data columns
        for src_col, dest_col in col_map.items():
            val = row_data.get(src_col)
            if pd.isna(val) if not isinstance(val, str) else False:
                val = None
            cell = ws_review.cell(row=data_row_idx, column=dest_col, value=val)
            cell.alignment = wrap_top
            cell.font = (
                hyperlink_font
                if isinstance(val, str) and val.startswith("=HYPERLINK(")
                else normal_font
            )

        ws_review.row_dimensions[data_row_idx].height = 15

    # Hide columns
    for col_idx in REVIEW_HIDDEN_COLS:
        ws_review.column_dimensions[get_column_letter(col_idx)].hidden = True

    # Autofilter on data range
    if len(df_needs_review) > 0:
        ws_review.auto_filter.ref = f"A1:{get_column_letter(len(REVIEW_HEADERS))}{len(df_needs_review) + 1}"

    # --- "Universe of Valid PINs" sheet ---
    ws_pins = wb.create_sheet("Universe of Valid PINs")
    ws_pins.cell(row=1, column=1, value="pin").font = bold_font

    # --- Save as .xlsm, transplanting VBA + Power Query from the demo ---
    # openpyxl cannot write .xlsm natively, so we:
    #   1. Save the workbook to a temp .xlsx
    #   2. Re-pack it as .xlsm, injecting binary/XML assets from the demo file
    import zipfile as _zf

    tmp_xlsx = file_name_review.replace(".xlsm", "_tmp.xlsx")
    wb.save(tmp_xlsx)

    # The demo .xlsm lives alongside this script in a "templates" folder.
    _script_dir = os.path.dirname(os.path.abspath(__file__))
    _demo_xlsm = os.path.join(
        _script_dir, "templates", "permits_needing_review.xlsm"
    )

    # Files transplanted verbatim from the demo into the new workbook
    _transplant = [
        "xl/vbaProject.bin",
        "xl/connections.xml",
        "xl/queryTables/queryTable1.xml",
        "xl/tables/_rels/table1.xml.rels",
        "customXml/item1.xml",
        "customXml/itemProps1.xml",
        "customXml/_rels/item1.xml.rels",
    ]

    with (
        _zf.ZipFile(tmp_xlsx, "r") as zin,
        _zf.ZipFile(file_name_review, "w", _zf.ZIP_DEFLATED) as zout,
        _zf.ZipFile(_demo_xlsm, "r") as zdemo,
    ):
        demo_names = set(zdemo.namelist())

        # Patch [Content_Types].xml to register vbaProject.bin and set xlsm type
        ct_xml = zin.read("[Content_Types].xml").decode("utf-8")
        if "vbaProject.bin" not in ct_xml:
            ct_xml = ct_xml.replace(
                "</Types>",
                '<Override PartName="/xl/vbaProject.bin" '
                'ContentType="application/vnd.ms-office.activeX+xml"/>'
                '<Default Extension="bin" '
                'ContentType="application/vnd.ms-office.activeX"/>'
                "</Types>",
            )
        ct_xml = ct_xml.replace(
            'ContentType="application/vnd.openxmlformats-officedocument'
            '.spreadsheetml.sheet.main+xml"',
            'ContentType="application/vnd.ms-excel.sheet.macroEnabled.main+xml"',
        )
        zout.writestr("[Content_Types].xml", ct_xml)

        # Patch xl/_rels/workbook.xml.rels to declare the VBA project relationship
        rels_xml = zin.read("xl/_rels/workbook.xml.rels").decode("utf-8")
        if "vbaProject.bin" not in rels_xml:
            rels_xml = rels_xml.replace(
                "</Relationships>",
                '<Relationship Id="rIdVBA" '
                'Type="http://schemas.microsoft.com/office/2006/'
                'relationships/vbaProject" '
                'Target="vbaProject.bin"/>'
                "</Relationships>",
            )
        zout.writestr("xl/_rels/workbook.xml.rels", rels_xml)

        # Copy everything else from the new xlsx unchanged
        _skip = {"[Content_Types].xml", "xl/_rels/workbook.xml.rels"} | set(
            _transplant
        )
        for item in zin.namelist():
            if item not in _skip:
                zout.writestr(item, zin.read(item))

        # Inject the transplanted assets from the demo
        for name in _transplant:
            if name in demo_names:
                zout.writestr(name, zdemo.read(name))

    os.remove(tmp_xlsx)
    print(f"Saved review workbook (.xlsm) to {file_name_review}")


if __name__ == "__main__":
    # Parse command line arguments
    start_date, end_date, deduplicate = parse_args()

    # Set up database connection cursor to query Athena
    conn = connect(
        s3_staging_dir=os.getenv(
            "AWS_ATHENA_S3_STAGING_DIR",
            "s3://ccao-athena-results-us-east-1",
        ),
        region_name=os.getenv(
            "AWS_REGION",
            "us-east-1",
        ),
    )
    cursor = conn.cursor()

    # Query for the current assessment year, which we will use to build a
    # universe of all current PINs to use for validating permit PINs. Smartfile
    # validates PINs against the current assessment year, not the date of
    # permit issue, so we need to match that logic
    print("Querying for current assessment year")
    year = get_current_assessment_year(cursor)

    pin_cache_filename = get_pin_cache_filename(year)
    if os.path.exists(pin_cache_filename):
        print(f"Loading Chicago PIN universe data from {pin_cache_filename}")
        chicago_pin_universe = pd.read_csv(
            pin_cache_filename,
            dtype={"pin": "string", "pin10": "string"},
        )
    else:
        print(f"Pulling {year} PINs from Athena")
        chicago_pin_universe = pull_existing_pins_from_athena(cursor, year)

    print(f"Downloading permits between {start_date} and {end_date}")
    permits = download_permits(start_date, end_date)
    print(f"Cleaning {len(permits)} permit{'' if len(permits) == 1 else 's'}")

    # Chicago permit data does not include city and state, but smartfile
    # expects it, so add it manually
    permits["city_state"] = "CHICAGO, IL"

    permits_expanded = expand_multi_pin_permits(permits)

    permits_pin = format_pin(permits_expanded)

    permits_renamed = organize_columns(permits_expanded)

    permits_validated = flag_invalid_pins(
        permits_renamed, chicago_pin_universe
    )

    joined_permits = add_address_link_and_suggested_pins(
        permits_validated, chicago_pin_universe
    )

    permits_shortened = flag_fix_long_fields(joined_permits)

    if deduplicate:
        print(
            "Number of permits prior to deduplication: "
            f"{len(permits_shortened)}"
        )
        permits_deduped = deduplicate_permits(
            cursor, permits_shortened, start_date, end_date
        )
        print(f"Number of permits after deduplication: {len(permits_deduped)}")
    else:
        permits_deduped = permits_shortened

    file_base_name = gen_file_base_name(start_date, end_date)

    save_xlsx_files(permits_deduped, 200, file_base_name)
