# This script is now dated. It was used to review permits based on a different set of techniques.
# It is now saved for reference sake.
import argparse  # noqa: I001
import os
from datetime import datetime

import openpyxl
import pandas as pd
from helper import (
    filled_columns,
    finalize_columns,
    get_pin_cache_filename,
    normalize_pin,
    pull_existing_pins_from_athena,
    required_columns,
)
from openpyxl.utils.datetime import from_excel
from pyathena import connect

conn = connect(
    s3_staging_dir=os.getenv(
        "AWS_ATHENA_S3_STAGING_DIR",
        "s3://ccao-athena-results-us-east-1",
    ),
    region_name=os.getenv(
        "AWS_REGION",
        "us-east-1",
    ),
)
cursor = conn.cursor()
FLAG_FILL_COLORS = {
    "FFFFFF00",  # yellow (ARGB)
    "FFFFC000",  # orange (ARGB)
    # For some reason, one color is not recognized with the hex, but
    # only with the theme value.
    ("theme", 7, 0.3999755851924192),
}


def pin_cell_matches_flag(pin_cell) -> bool:
    """Return True if the PIN cell has a background color in FLAG_FILL_COLORS."""
    if pin_cell is None:
        return False

    fg = getattr(pin_cell.fill, "fgColor", None)
    if fg is None:
        return False

    #  RGB Fills
    rgb = getattr(fg, "rgb", None)
    if rgb:
        val = str(rgb).upper().lstrip("#")
        if val in FLAG_FILL_COLORS:
            return True
        if len(val) == 8 and val[2:] in FLAG_FILL_COLORS:
            return True

    # Theme-based fills (there is one color which is theme-based)
    if getattr(fg, "type", None) == "theme":
        theme = getattr(fg, "theme", None)
        tint = getattr(fg, "tint", None)

        # normalize tint to avoid float precision issues
        if isinstance(tint, float):
            tint = round(tint, 6)

        # also normalize the tuples stored in FLAG_FILL_COLORS
        # Tried without these steps and it wasn't caught.
        normalized_flag_colors = {
            (
                t[0],
                t[1],
                round(t[2], 6),
            )
            if isinstance(t, tuple) and len(t) == 3 and isinstance(t[2], float)
            else t
            for t in FLAG_FILL_COLORS
        }

        if ("theme", theme, tint) in normalized_flag_colors:
            return True

    return False


def remove_flagged_rows_from_original_xlsx(
    file_path: str, cleaned_xlsx_path: str, sheet_name: str
) -> str:
    """
    Remove any rows in the given sheet whose 'PIN* [PARID]' cell fill color
    matches FLAG_FILL_COLORS, and save a copy of the workbook preserving formatting.
    This is mainly for QC'ing to make sure that colors meant to be preserved
    are not accidentally deleted from the data. Moving forward it should not be important.
    """
    wb = openpyxl.load_workbook(file_path, data_only=False)
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"Sheet '{sheet_name}' not found in workbook.")
    ws = wb[sheet_name]

    # Read header row directly from worksheet cells
    header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=False))
    original_header = [
        c.value if c.value is not None else "" for c in header_cells
    ]
    header_index = {col: i for i, col in enumerate(original_header)}
    pin_idx = header_index.get("PIN* [PARID]")

    # Determine Excel row numbers to delete
    rows_to_delete = []
    for r in range(2, ws.max_row + 1):
        pin_cell = ws.cell(row=r, column=pin_idx + 1)
        if pin_cell_matches_flag(pin_cell):
            rows_to_delete.append(r)

    # Delete bottom-up so row indices don't shift
    for r in reversed(rows_to_delete):
        ws.delete_rows(r, 1)

    wb.save(cleaned_xlsx_path)
    return cleaned_xlsx_path


def extract_flagged_rows_from_sheet(
    wb: openpyxl.Workbook, sheet_name: str
) -> tuple[pd.DataFrame, str, str]:
    """
    Read a sheet, keep only rows where PIN cell is flagged, and return:
    (df_flagged_only, start_date_yyyy_mm_dd, end_date_yyyy_mm_dd)
    """
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"Sheet '{sheet_name}' not found in workbook.")
    ws = wb[sheet_name]

    rows_values = list(ws.iter_rows(values_only=True))
    rows_cells = list(ws.iter_rows(values_only=False))

    # Normalize header row
    original_header = [c if c is not None else "" for c in rows_values[0]]
    header_index = {col: i for i, col in enumerate(original_header)}
    pin_idx = header_index.get("PIN* [PARID]")

    issue_col = "Issue Date* [PERMDT]"
    issue_idx = header_index.get(issue_col)
    if issue_idx is None:
        raise ValueError(
            f"Column '{issue_col}' not found in '{sheet_name}' sheet."
        )

    issue_dates = []
    for row in rows_values[1:]:
        raw = row[issue_idx] if issue_idx < len(row) else None
        if raw is None or raw == "":
            continue

        if isinstance(raw, datetime):
            dt = raw
        elif isinstance(raw, (int, float)):
            dt = from_excel(raw)
        else:
            dt = datetime.strptime(str(raw).strip(), "%m/%d/%Y")

        issue_dates.append(dt)

    if not issue_dates:
        start_date = None
        end_date = None
    else:
        start_date = min(issue_dates).strftime("%Y-%m-%d")
        end_date = max(issue_dates).strftime("%Y-%m-%d")

    # Only keep rows where the PIN cell background matches the flag colors
    flagged_rows = []

    for row_vals, row_cells in zip(rows_values[1:], rows_cells[1:]):
        row_vals = list(row_vals)

        # Extract PIN background cell
        pin_cell = None
        if pin_idx is not None and pin_idx < len(row_cells):
            pin_cell = row_cells[pin_idx]

        if not pin_cell_matches_flag(pin_cell):
            continue

        # Retain ALL columns from the sheet for review/QC
        new_row = {}
        for col in original_header:
            idx = header_index.get(col)
            if idx is not None and idx < len(row_vals):
                val = row_vals[idx]
                new_row[col] = "" if val is None else val
            else:
                new_row[col] = ""

        # Normalize PIN for upload processing
        if "PIN* [PARID]" in new_row:
            new_row["PIN* [PARID]"] = normalize_pin(
                str(new_row["PIN* [PARID]"])
            )

        flagged_rows.append(new_row)

    df_flagged_only = pd.DataFrame(flagged_rows)
    return df_flagged_only, start_date, end_date


def write_outputs_for_sheet(
    file_path: str,
    sheet_name: str,
    chicago_pin_universe: pd.DataFrame,
    start_date: str,
    end_date: str,
) -> None:
    # Upload batching setup
    batch_size = 250
    batch_number = 1

    wb = openpyxl.load_workbook(file_path, data_only=True)
    df_flagged_only, s, e = extract_flagged_rows_from_sheet(wb, sheet_name)

    # Prefer the precomputed cache range; fall back to sheet-derived range
    start_date = start_date or s
    end_date = end_date or e

    if start_date is None or end_date is None:
        raise ValueError(
            f"Could not determine start/end issue dates for sheet '{sheet_name}'."
        )

    # New: separate folder per sheet, but suffix depends on output type
    safe_sheet = sheet_name.lower().replace(" ", "_")
    base_prefix = f"{start_date}_to_{end_date}_{safe_sheet}_permits"

    upload_folder = f"{base_prefix}_ready_for_upload"
    need_review_folder = f"{base_prefix}_need_review"
    color_testing_folder = f"{base_prefix}_color_testing"

    os.makedirs(upload_folder, exist_ok=True)
    os.makedirs(need_review_folder, exist_ok=True)
    os.makedirs(color_testing_folder, exist_ok=True)

    out = finalize_columns(
        df_flagged_only, filled_columns, chicago_pin_universe
    )
    upload_df = out["upload"].copy()
    upload_df["Issue Date* [PERMDT]"] = pd.to_datetime(
        upload_df["Issue Date* [PERMDT]"], errors="coerce"
    ).dt.strftime("%m/%d/%Y")

    need_review_df = out["need_review"].copy()
    need_review_df["Issue Date* [PERMDT]"] = pd.to_datetime(
        need_review_df["Issue Date* [PERMDT]"], errors="coerce"
    ).dt.strftime("%m/%d/%Y")

    # write need_review as XLSX (retain all fields, including validation flags)
    need_review_path = os.path.join(
        need_review_folder, f"{base_prefix}_need_review.xlsx"
    )
    need_review_df["# [LLINE]"] = range(1, len(need_review_df) + 1)

    # Order columns: required first, then everything else (extras + validation flags)
    # This is to help for QC by res-val following our output.
    ordered_review_cols = required_columns + [
        c for c in need_review_df.columns if c not in required_columns
    ]
    need_review_df = need_review_df.reindex(columns=ordered_review_cols)

    need_review_df.to_excel(
        need_review_path,
        index=False,
        engine="openpyxl",
    )
    print(f"[{sheet_name}] Need-review XLSX saved to: {need_review_path}")
    print(
        f"[{sheet_name}] Total need-review rows written: {len(need_review_df)}"
    )

    # write upload rows in batches as XLSX (keep required schema only)
    for start in range(0, len(upload_df), batch_size):
        batch = upload_df.iloc[start : start + batch_size].copy()
        batch["# [LLINE]"] = range(1, len(batch) + 1)

        # Ensure column order
        batch = batch.reindex(columns=required_columns)

        upload_batch_path = os.path.join(
            upload_folder,
            f"{base_prefix}_ready_for_upload_{batch_number}.xlsx",
        )
        batch.to_excel(
            upload_batch_path,
            index=False,
            engine="openpyxl",
        )

        print(f"[{sheet_name}] Upload batch saved to: {upload_batch_path}")

        if start + batch_size < len(upload_df):
            batch_number += 1

    # Keep the cleaned workbook in the color_testing folder
    cleaned_xlsx_path = os.path.join(
        color_testing_folder,
        f"{base_prefix}_color_testing_cleaned_flagged_rows_removed.xlsx",
    )
    remove_flagged_rows_from_original_xlsx(
        file_path=file_path,
        cleaned_xlsx_path=cleaned_xlsx_path,
        sheet_name=sheet_name,
    )
    print(
        f"[{sheet_name}] Workbook with flagged rows removed saved to: {cleaned_xlsx_path}"
    )


def format_reviewed_permits_for_upload(file_path: str) -> None:
    # Load once to compute overall date range for cache
    wb_dates = openpyxl.load_workbook(file_path, data_only=True)

    sheets_to_process = ["PIN Errors", "Other Errors"]

    all_dates = []
    for sheet in sheets_to_process:
        _, s, e = extract_flagged_rows_from_sheet(wb_dates, sheet)
        if s and e:
            all_dates.append(datetime.strptime(s, "%Y-%m-%d"))
            all_dates.append(datetime.strptime(e, "%Y-%m-%d"))

    if not all_dates:
        raise ValueError(
            "Could not determine issue date ranges from the requested sheets."
        )

    start_date = min(all_dates).strftime("%Y-%m-%d")
    end_date = max(all_dates).strftime("%Y-%m-%d")

    pin_cache_filename = get_pin_cache_filename(start_date, end_date)
    if os.path.exists(pin_cache_filename):
        print(f"Loading Chicago PIN universe data from {pin_cache_filename}")
        chicago_pin_universe = pd.read_csv(
            pin_cache_filename,
            dtype={"pin": "string", "pin10": "string"},
        )
    else:
        print("Pulling PINs from Athena")
        chicago_pin_universe = pull_existing_pins_from_athena(
            cursor, start_date, end_date
        )
        chicago_pin_universe.to_csv(
            pin_cache_filename, index=False, encoding="utf-8"
        )
        print(f"Saved Chicago PIN universe data to {pin_cache_filename}")

    # Process each sheet into its own output folders
    for sheet_name in sheets_to_process:
        write_outputs_for_sheet(
            file_path=file_path,
            sheet_name=sheet_name,
            chicago_pin_universe=chicago_pin_universe,
            start_date=start_date,
            end_date=end_date,
        )


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("file_path", help="Path to the Excel file")
    args = parser.parse_args()
    format_reviewed_permits_for_upload(args.file_path)
